import streamlit as st
import pandas as pd
import json
import os
from pathlib import Path
import io
import re
from openpyxl import load_workbook

st.set_page_config(
    page_title="智能结算单汇总工具",
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.markdown("""
<style>
    .main .block-container { padding-top: 2rem; padding-bottom: 2rem; }
    .stButton>button { background-color: #4c85e4; color: white; border-radius: 6px; padding: 0.5rem 1rem; border: none; }
    .stButton>button:hover { background-color: #3a6fc4; }
</style>
""", unsafe_allow_html=True)

DEFAULT_APP_BUSINESS_LINES = [
    {"app_name": "Novellair", "business_line": "付费网文"},
    {"app_name": "Novelbar", "business_line": "付费网文"},
    {"app_name": "Literie", "business_line": "付费网文"},
    {"app_name": "Metronovel", "business_line": "付费网文"},
    {"app_name": "Snackread", "business_line": "付费网文"},
    
    {"app_name": "Minishorts", "business_line": "付费短剧"},
    {"app_name": "Mintime", "business_line": "付费短剧"},
    {"app_name": "Minitime", "business_line": "付费短剧"},
]

DEFAULT_EXPORT_COLUMNS = [
    '结算单确认日期', '主体', '业务线', 'APP名称', '收入类型', 
    '支付渠道', '结算周期', '结算金额（含税）_新币', '服务费_费率'
]

if 'config' not in st.session_state:
    st.session_state.config = {
        "app_business_lines": [],
        "export_columns": DEFAULT_EXPORT_COLUMNS.copy()
    }
if 'extracted_data' not in st.session_state:
    st.session_state.extracted_data = pd.DataFrame()

def parse_excel_to_df(file_obj):
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))
    df = pd.DataFrame(data)
    return df

def find_header_row(df):
    header_keywords = ['项目', '类型', '周期', '金额', '日期', '渠道', '收入']
    for i in range(min(10, len(df))):
        row_values = df.iloc[i].values
        found_keywords = sum(1 for v in row_values if pd.notna(v) and any(kw in str(v) for kw in header_keywords))
        non_empty_count = sum(1 for v in row_values if pd.notna(v) and str(v).strip())
        if found_keywords >= 2 and non_empty_count >= 3:
            return i
    return 0

def get_data_rows(df):
    header_row = find_header_row(df)
    summary_keywords = ['合计', '总计', '小计', '合计金额', '备注', '签字', '确认', '审核', '甲方', '乙方', '服务费']
    data_rows = []
    for idx in range(header_row + 1, len(df)):
        row_values = df.iloc[idx].values
        non_empty_count = sum(1 for v in row_values if pd.notna(v) and str(v).strip())
        if non_empty_count == 0:
            continue
        
        is_summary = False
        
        # 检查是否是汇总行或元信息行
        for v in row_values:
            if pd.notna(v):
                v_str = str(v).strip()
                if any(kw in v_str for kw in summary_keywords):
                    is_summary = True
                    break
        
        # 检查第一个单元格：如果没有项目相关内容，可能是汇总行
        first_value = row_values[0] if len(row_values) > 0 else None
        if first_value is not None and pd.notna(first_value):
            first_str = str(first_value).strip()
            # 检查是否包含项目特征：内容分成、广告等，或者包含APP名称（包含大写字母）
            has_project_feature = any(kw in first_str for kw in ['内容分成', '广告', '充值', '订阅', '内购']) or any(c.isupper() for c in first_str)
            if not has_project_feature:
                is_summary = True
        
        if is_summary:
            continue
        
        data_rows.append(row_values)
    return data_rows

def extract_meta_info(df):
    meta = {'主体': '', '结算单确认日期': '', '服务费_费率': ''}
    for idx in reversed(range(len(df))):
        row_values = df.iloc[idx].values
        for v in row_values:
            if pd.notna(v):
                v_str = str(v).strip()
                if '甲方' in v_str or 'Party A' in v_str:
                    party_a_match = re.search(r'[：:]\s*(.*)', v_str)
                    if party_a_match:
                        meta['主体'] = party_a_match.group(1).strip()
                    else:
                        for v2 in row_values:
                            if pd.notna(v2) and str(v2).strip() and '甲方' not in str(v2):
                                meta['主体'] = str(v2).strip()
                                break
        for v in row_values:
            if pd.notna(v):
                v_str = str(v).strip()
                if ('日期' in v_str or 'Date' in v_str) and not meta['结算单确认日期']:
                    date_match = re.search(r'(\d{4}[-/年]\s*\d{1,2}[-/月]\s*\d{1,2}[日]?)', v_str)
                    if date_match:
                        meta['结算单确认日期'] = date_match.group(1).replace('年', '-').replace('月', '-').replace('日', '')
                    else:
                        for v2 in row_values:
                            if pd.notna(v2):
                                v2_str = str(v2).strip()
                                date_match2 = re.search(r'(\d{4}[-/]\s*\d{1,2}[-/]\s*\d{1,2})', v2_str)
                                if date_match2:
                                    meta['结算单确认日期'] = date_match2.group(1)
                                    break
        for v in row_values:
            if pd.notna(v):
                v_str = str(v).strip()
                if '服务费' in v_str and not meta['服务费_费率']:
                    fee_match = re.search(r'([\d\.%]+)', v_str)
                    if fee_match:
                        meta['服务费_费率'] = fee_match.group(1)
                    else:
                        for v2 in row_values:
                            if pd.notna(v2):
                                v2_str = str(v2).strip()
                                fee_match2 = re.search(r'([\d\.%]+)', v2_str)
                                if fee_match2:
                                    meta['服务费_费率'] = fee_match2.group(1)
                                    break
    return meta

def split_project_column(project_text):
    if pd.isna(project_text):
        return None, None, None
    text = str(project_text).strip()
    type_keywords = ['内容分成', 'Mintegral 广告', 'liftoff广告', '广告', '充值', '订阅', '内购']
    type_keywords_sorted = sorted(type_keywords, key=lambda x: len(x), reverse=True)
    for kw in type_keywords_sorted:
        if kw in text:
            income_type = kw
            if '广告' in income_type:
                income_type = '广告'
            app_part = text[:text.index(kw)].strip()
            
            if '广告' in kw:
                words = app_part.split()
                first_word = words[0] if words else ''
                return first_word, first_word, income_type
            
            first_app = app_part
            
            if '、' in app_part:
                first_app = app_part.split('、')[0].strip()
            elif '+' in app_part:
                first_app = app_part.split('+')[0].strip()
            elif '，' in app_part:
                first_app = app_part.split('，')[0].strip()
            elif ',' in app_part:
                first_app = app_part.split(',')[0].strip()
            else:
                first_app = app_part
            
            first_app = first_app.strip()
            
            return app_part, first_app, income_type
    
    parts = re.split(r'[+、，,]', text)
    if len(parts) >= 2:
        return text, parts[0].strip(), ''
    return text, text, ''

def get_business_line(app_name, app_business_lines):
    if not app_name:
        return ''
    for mapping in app_business_lines:
        if mapping.get('app_name') == app_name:
            return mapping.get('business_line', '')
    for mapping in DEFAULT_APP_BUSINESS_LINES:
        if mapping.get('app_name') == app_name:
            return mapping.get('business_line', '')
    return ''

def process_settlement_file(file_obj, app_business_lines):
    df_raw = parse_excel_to_df(file_obj)
    meta = extract_meta_info(df_raw)
    
    header_row = find_header_row(df_raw)
    headers = df_raw.iloc[header_row].values
    headers = [str(h).strip() if pd.notna(h) else f"列{i}" for i, h in enumerate(headers)]
    
    data_rows = get_data_rows(df_raw)
    df_data = pd.DataFrame(data_rows, columns=headers)
    
    results = []
    project_col = None
    amount_col = None
    period_col = None
    type_col = None
    
    for col in headers:
        if '项目' in col:
            project_col = col
        if '金额' in col:
            amount_col = col
        if '周期' in col or ('日期' in col and not period_col):
            period_col = col
        if '类型' in col or '渠道' in col:
            type_col = col
    
    for _, row in df_data.iterrows():
        result = {}
        result.update(meta)
        
        if project_col and pd.notna(row.get(project_col)):
            app_display, app_match, income_type = split_project_column(row.get(project_col))
            result['APP名称'] = app_display
            result['收入类型'] = income_type
            result['业务线'] = get_business_line(app_match, app_business_lines)
        
        if type_col and pd.notna(row.get(type_col)):
            result['支付渠道'] = row.get(type_col)
        if period_col and pd.notna(row.get(period_col)):
            result['结算周期'] = row.get(period_col)
        
        if amount_col and pd.notna(row.get(amount_col)):
            try:
                result['结算金额（含税）_新币'] = float(row.get(amount_col))
            except (ValueError, TypeError):
                result['结算金额（含税）_新币'] = row.get(amount_col)
        
        results.append(result)
    return results

def aggregate_results(results):
    if not results:
        return []
    df = pd.DataFrame(results)
    required_cols = ['APP名称', '收入类型', '支付渠道', '结算周期', '结算金额（含税）_新币']
    for col in required_cols:
        if col not in df.columns:
            df[col] = None
    if '业务线' not in df.columns:
        df['业务线'] = ''
    if '服务费_费率' not in df.columns:
        df['服务费_费率'] = ''
    group_cols = ['主体', '结算单确认日期', 'APP名称', '业务线', '收入类型', '支付渠道', '结算周期', '服务费_费率']
    group_cols = [col for col in group_cols if col in df.columns]
    aggregated = df.groupby(group_cols, dropna=False).agg({
        '结算金额（含税）_新币': 'sum'
    }).reset_index()
    final_cols = ['主体', '结算单确认日期', 'APP名称', '业务线', '收入类型', '支付渠道', '结算周期', '服务费_费率', '结算金额（含税）_新币']
    for col in final_cols:
        if col not in aggregated.columns:
            aggregated[col] = None
    return aggregated[final_cols].to_dict('records')

st.title("🤖 智能结算单汇总工具")
st.markdown("---")

tab1, tab2, tab3 = st.tabs(["📊 数据提取", "🔍 预览与验证", "🔧 配置管理"])

with tab1:
    st.header("上传并提取结算单数据")
    st.info("将自动提取：主体、结算单确认日期、APP名称、业务线、收入类型、支付渠道、结算周期、服务费_费率、结算金额（含税）_新币")
    
    uploaded_files = st.file_uploader(
        "上传结算单 Excel 文件（可多选）",
        type=['xlsx', 'xls'],
        accept_multiple_files=True
    )
    
    # 检查是否有提取的数据
    if 'extracted_data' not in st.session_state:
        st.session_state.extracted_data = None
    
    # 提取数据按钮
    if uploaded_files and st.button("开始提取数据", type="primary"):
        app_business_lines = st.session_state.config.get("app_business_lines", [])
        all_results = []
        for file in uploaded_files:
            try:
                file.seek(0)
                results = process_settlement_file(file, app_business_lines)
                if results:
                    all_results.extend(results)
                    st.success(f"✅ 成功处理文件：{file.name}")
                else:
                    st.warning(f"⚠️ 文件 {file.name} 中没有找到数据行")
            except Exception as e:
                st.error(f"❌ 处理文件 {file.name} 出错：{str(e)}")
                import traceback
                st.text(traceback.format_exc())
        
        if all_results:
            final_results = aggregate_results(all_results)
            final_df = pd.DataFrame(final_results)
            st.session_state.extracted_data = final_df
    
    # 如果有提取的数据，显示编辑和导出界面
    if st.session_state.extracted_data is not None:
        st.markdown("---")
        st.subheader("🎉 提取完成！（已按相同条件汇总）")
        
        edited_df = st.data_editor(
            st.session_state.extracted_data,
            num_rows="dynamic",
            use_container_width=True
        )
        
        st.markdown("---")
        st.subheader("⚙️ 导出设置")
        
        current_cols = st.session_state.config.get("export_columns", DEFAULT_EXPORT_COLUMNS.copy())
        available_cols = ['结算单确认日期', '主体', '业务线', 'APP名称', '收入类型', '支付渠道', '结算周期', '结算金额（含税）_新币', '服务费_费率']
        
        selected_cols = st.multiselect(
            "选择导出列（按顺序）",
            options=available_cols,
            default=current_cols,
            format_func=lambda x: x,
            key="export_cols_select"
        )
        
        if st.button("💾 保存导出列顺序", key="save_export_cols"):
            st.session_state.config["export_columns"] = selected_cols
            st.success("✅ 导出列顺序已保存！")
        
        if st.session_state.config.get("export_columns", []) != current_cols:
            st.warning("⚠️ 当前导出列顺序与保存的不同，请点击上方按钮保存")
        
        # 确保列存在
        valid_cols = [col for col in selected_cols if col in edited_df.columns]
        ordered_df = edited_df[valid_cols] if valid_cols else edited_df
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            ordered_df.to_excel(writer, index=False, sheet_name='汇总数据')
        
        st.download_button(
            label="📥 下载汇总 Excel",
            data=output.getvalue(),
            file_name="结算单汇总.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_excel"
        )

with tab2:
    st.header("单文件预览与验证")
    st.subheader("上传一个文件查看提取效果")
    
    sample_file = st.file_uploader("上传一个结算单文件预览提取验证", type=['xlsx', 'xls'], key="preview_sample")
    
    if sample_file:
        sample_file.seek(0)
        df_raw = parse_excel_to_df(sample_file)
        
        st.markdown("### 📋 Excel 原始内容（前20行）")
        display_rows = min(20, len(df_raw))
        display_df = df_raw.head(display_rows)
        display_df.columns = [f"列{i}" for i in range(len(display_df.columns))]
        display_df.index = [f"行{i+1}" for i in range(display_rows)]
        st.dataframe(display_df, use_container_width=True)
        
        st.markdown("---")
        st.markdown("### ✅ 提取结果")
        
        try:
            app_business_lines = st.session_state.config.get("app_business_lines", [])
            results = process_settlement_file(sample_file, app_business_lines)
            if results:
                st.markdown("#### 📝 提取到的元信息：")
                meta_col1, meta_col2 = st.columns(2)
                with meta_col1:
                    st.info(f"**主体**：{results[0].get('主体', '未找到')}")
                with meta_col2:
                    st.info(f"**结算单确认日期**：{results[0].get('结算单确认日期', '未找到')}")
                
                final_results = aggregate_results(results)
                final_df = pd.DataFrame(final_results)
                
                st.markdown("#### 📊 提取的数据（已汇总）：")
                st.dataframe(final_df, use_container_width=True)
            else:
                st.warning("未找到数据行")
        except Exception as e:
            st.error(f"提取出错：{str(e)}")
            import traceback
            st.text(traceback.format_exc())

with tab3:
    st.header("🔧 配置管理")
    
    st.subheader("📋 App-业务线关联关系")
    st.info("在这里维护 App 名称和业务线的关联关系，提取数据时会自动匹配")
    
    app_business_lines = st.session_state.config.get("app_business_lines", [])
    
    with st.expander("📋 默认配置（固定不会变）", expanded=True):
        default_df = pd.DataFrame(DEFAULT_APP_BUSINESS_LINES)
        default_df.columns = ["App 名称", "业务线"]
        st.dataframe(default_df, use_container_width=True)
    
    st.markdown("---")
    
    st.subheader("➕ 新增临时关联关系")
    col1, col2, col3 = st.columns(3)
    with col1:
        new_app = st.text_input("新增 App 名称")
    with col2:
        new_business_line = st.text_input("新增业务线（例如：付费网文）")
    with col3:
        st.text("")
        if st.button("添加到临时配置"):
            if new_app and new_business_line:
                in_default = any(m["app_name"] == new_app for m in DEFAULT_APP_BUSINESS_LINES)
                if in_default:
                    st.warning(f"⚠️ App \"{new_app}\" 已在默认配置中")
                else:
                    exists = any(m["app_name"] == new_app for m in app_business_lines)
                    if exists:
                        st.warning(f"⚠️ App \"{new_app}\" 已存在临时配置中")
                    else:
                        app_business_lines.append({
                            "app_name": new_app,
                            "business_line": new_business_line
                        })
                        st.session_state.config["app_business_lines"] = app_business_lines
                        st.success(f"✅ 添加成功！{new_app} → {new_business_line}")
            else:
                st.warning("⚠️ 请填写 App 名称和业务线")
    
    st.markdown("---")
    
    st.subheader("📝 临时关联关系（可添加和删除）")
    
    if app_business_lines:
        df_mappings = pd.DataFrame(app_business_lines)
        df_mappings.columns = ["App 名称", "业务线"]
        
        edited_mappings = st.data_editor(
            df_mappings,
            num_rows="dynamic",
            use_container_width=True
        )
        
        if st.button("💾 保存临时配置修改"):
            updated_mappings = []
            for _, row in edited_mappings.iterrows():
                if pd.notna(row["App 名称"]) and pd.notna(row["业务线"]):
                    updated_mappings.append({
                        "app_name": str(row["App 名称"]),
                        "business_line": str(row["业务线"])
                    })
            st.session_state.config["app_business_lines"] = updated_mappings
            st.success("✅ 保存成功！")
    else:
        st.info("暂无临时配置")
    
    st.markdown("---")
    st.subheader("📦 配置导入/导出")
    
    col_export, col_import, col_reset = st.columns(3)
    with col_export:
        all_mappings = DEFAULT_APP_BUSINESS_LINES.copy()
        existing_apps = set(m["app_name"] for m in all_mappings)
        for m in app_business_lines:
            if m["app_name"] not in existing_apps:
                all_mappings.append(m)
        
        config_to_export = {
            "app_business_lines": all_mappings,
            "export_columns": st.session_state.config.get("export_columns", DEFAULT_EXPORT_COLUMNS.copy())
        }
        config_json = json.dumps(config_to_export, ensure_ascii=False, indent=2)
        st.download_button(
            label="📤 导出完整配置",
            data=config_json,
            file_name="settlement_config.json",
            mime="application/json"
        )
    
    with col_import:
        config_file = st.file_uploader(
            "📥 导入配置文件",
            type=['json'],
            key="config_import"
        )
        if config_file:
            try:
                config_data = json.load(config_file)
                if "app_business_lines" in config_data:
                    default_apps = set(m["app_name"] for m in DEFAULT_APP_BUSINESS_LINES)
                    imported_mappings = config_data["app_business_lines"]
                    new_mappings = [m for m in imported_mappings if m["app_name"] not in default_apps]
                    st.session_state.config["app_business_lines"] = new_mappings
                if "export_columns" in config_data:
                    st.session_state.config["export_columns"] = config_data["export_columns"]
                st.success("✅ 配置导入成功！")
            except Exception as e:
                st.error(f"❌ 导入配置失败：{str(e)}")
    
    with col_reset:
        if st.button("🔄 重置配置"):
            st.session_state.config = {
                "app_business_lines": [],
                "export_columns": DEFAULT_EXPORT_COLUMNS.copy()
            }
            st.success("✅ 配置已重置！")

st.markdown("---")
st.info("💡 提示：使用「配置导入/导出」功能来保存和分享您的配置！")